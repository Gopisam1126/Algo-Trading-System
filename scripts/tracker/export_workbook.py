"""Regenerate tracker_data.json from the live workbook.

The artifact's HTML template is fixed; only the payload changes. Each sheet has
its header on a different row, so the header is FOUND rather than assumed —
guessing it silently produces a table of empty strings that still renders.
"""

import datetime as dt
import json
import os

import openpyxl

#: Resolved from THIS file's location, never hard-coded. These scripts used
#: to live in a session scratchpad and were lost once when it was cleared;
#: anything the `sdlc` skill instructs has to survive that.
HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
XL = os.path.join(REPO, "Documents", "BACKLOG_Tracker.xlsx")
OUT = os.path.join(HERE, "tracker_data.json")

SHEETS = {
    "stories": ("Backlog", "Story ID"),
    "blockers": ("Blockers", "ID"),
    "qa": ("QA Results", None),
    "security": ("Security Findings", None),
    "sit": ("SIT Defects", "SIT ID"),
}


def clean(v):
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.isoformat()[:10]
    return str(v).strip()


def find_header(ws, marker):
    """Locate the header row. With a marker, match it in column A; without,
    take the first row that has 4+ non-empty cells and no formula."""
    for row in ws.iter_rows(min_row=1, max_row=12):
        values = [clean(c.value) for c in row]
        if marker is not None:
            if values and values[0] == marker:
                return row[0].row, values
        else:
            filled = [v for v in values if v]
            if len(filled) >= 4 and not any(v.startswith("=") for v in filled):
                return row[0].row, values
    raise AssertionError(f"no header found in {ws.title}")


def extract(ws, marker):
    header_row, header = find_header(ws, marker)

    # STOP at the first blank header column. The QA and Security sheets carry a
    # small summary pivot to the RIGHT of the data, separated by a gap, and it
    # reuses the same column names ("Result", "Severity"). Reading straight
    # across builds a dict where the pivot's copy overwrites the real column —
    # silently, because both keys are legitimate names and the value is simply
    # blank for newer rows.
    cols = []
    for i, h in enumerate(header):
        if not h:
            if cols:
                break
            continue
        if any(h == existing for _, existing in cols):
            continue
        cols.append((i, h))
    out = []
    for row in ws.iter_rows(min_row=header_row + 1):
        values = [clean(c.value) for c in row]
        if not values or not values[0]:
            continue
        if values[0].startswith("="):
            continue
        record = {h: (values[i] if i < len(values) else "") for i, h in cols}
        out.append(record)
    return header, out


wb = openpyxl.load_workbook(XL, data_only=False)
data = {}
for key, (sheet, marker) in SHEETS.items():
    header, rows = extract(wb[sheet], marker)
    data[key] = rows
    print(f"  {key:10} {len(rows):4} rows from {sheet!r}  cols={len(header)}")

# Sanity: the counts must look like the workbook, not like an empty parse.
assert len(data["stories"]) > 150, f"only {len(data['stories'])} stories parsed"
assert len(data["blockers"]) >= 8
assert len(data["qa"]) > 40
assert len(data["security"]) > 20
# `sit` may legitimately be empty — no SIT defects is a good day, not a
# parse failure. Only assert the sheet was found at all.
assert "sit" in data

path = OUT
with open(path, "w", encoding="utf-8", newline="\n") as fh:
    fh.write(json.dumps(data, ensure_ascii=False, indent=0))
print("written:", path)

from collections import Counter

print("\nstatus distribution:")
for k, v in Counter(s.get("Status", "") for s in data["stories"]).most_common():
    print(f"   {k or '(blank)':28} {v}")
print("\nblockers:")
for b in data["blockers"]:
    print(f"   {b.get('ID', '?'):6} {b.get('Status', '')[:44]}")
